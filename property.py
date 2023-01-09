
from openpyxl import load_workbook
from selenium import webdriver
import time
import sys
from datetime import date
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import threading
from selenium.common.exceptions import NoSuchElementException
import requests
import random
import datetime
import string
import os
from geopy import distance
from geopy.geocoders import Nominatim
import openpyxl

#start headless chrome
options = webdriver.ChromeOptions()
options.add_argument('headless')
driver = webdriver.Chrome(options=options)
geolocator = Nominatim(user_agent="user-agent")
googleApiKey= 'AIzaSyB2tQN6E8ujV_LslOAFU3ERgTgzO8dzKTc'
#jacksonville
username = 'thomassafar03@gmail.com'
password = 'eatit093'
pages=["https://www.crexi.com/lease/properties?types%5B%5D=Retail&placeIds%5B%5D=ChIJ66_O8Ra35YgR4sf8ljh9zcQ&sqFtMax=2300&sort=New%20Listings",
"https://www.crexi.com/lease/properties?types%5B%5D=Retail&placeIds%5B%5D=ChIJ66_O8Ra35YgR4sf8ljh9zcQ&sqFtMax=2300&sort=New%20Listings&page=2",
"https://www.crexi.com/lease/properties?types%5B%5D=Retail&placeIds%5B%5D=ChIJ66_O8Ra35YgR4sf8ljh9zcQ&sqFtMax=2300&page=3", 
"https://www.crexi.com/lease/properties?types%5B%5D=Retail&sqFtMax=2300&placeIds%5B%5D=ChIJw7zsvY7N5YgRP8kydmibdu0",
"https://www.crexi.com/lease/properties?types%5B%5D=Retail&sqFtMax=2300&placeIds%5B%5D=ChIJw7zsvY7N5YgRP8kydmibdu0&page=2",
"https://www.crexi.com/lease/properties?types%5B%5D=Retail&sqFtMax=2300&placeIds%5B%5D=ChIJkcQBieJI5IgRFfy4tQ2FGdw",
"https://www.crexi.com/lease/properties?types%5B%5D=Retail&sqFtMax=2300&placeIds%5B%5D=ChIJkcQBieJI5IgRFfy4tQ2FGdw&page=2",
]
#login

driver.get("https://www.crexi.com/lease/properties?types%5B%5D=Retail&placeIds%5B%5D=ChIJ66_O8Ra35YgR4sf8ljh9zcQ&sqFtMax=2300&sort=New%20Listings")
time.sleep(10)
#remove popup if it exists
try:
    driver.execute_script("document.getElementsByClassName('cui-modal-close ng-star-inserted')[0].click()")
except:
    pass
#click login
driver.execute_script('document.querySelector("body > crx-app > div > ng-component > crx-normal-page > div > crx-header > crx-header-content > div > div.transclude-auth.right-header-section > crx-logged-out-header > button").click()')
time.sleep(2)
#click login tab
driver.execute_script("document.getElementsByClassName('tab switch')[0].click()")
time.sleep(2)
emailBox = driver.execute_script('return document.querySelector("#login-form > div:nth-child(1) > label > input")')
emailBox.send_keys(username)
passwordBox = driver.execute_script('return document.querySelector("#login-form > div:nth-child(2) > label > input")')
passwordBox.send_keys(password)
time.sleep(2)
#finally click login
driver.execute_script('document.querySelector("#login-form > button").click()')
time.sleep(5)
i=0
links = []
for i in range(0, len(pages)):
    #check if 'Oh no! There aren’t any spaces that match your search. Remove filters or update filters to find more spaces:' not in document.body.innerHTML
    driver.get(pages[i])
    time.sleep(10)
    if 'Oh no! There aren’t any spaces that match your search. Remove filters or update filters to find more spaces:' in driver.execute_script('return document.body.innerHTML'):
        continue
    time.sleep(10)
    time.sleep(2)
    try:
        driver.execute_script('document.querySelector("#pagination-container > div > div > crx-select > crx-dropdown-button > div > div").click()')
        time.sleep(2)
        driver.execute_script('document.querySelector("#pagination-container > div > div > crx-select > crx-dropdown-button > div > crx-dropdown-portal > div > div > div.options > div:nth-child(5)").click()')
        time.sleep(2)
    except:
        print('cant change showing size')
        pass
    #save document.getElementsByClassName('cover-link')[0-99].href to list
    for i in range(0, 100):
        try:
            links.append(driver.find_elements(By.CLASS_NAME,'cover-link')[i].get_attribute('href'))
            print(links[i])
        except:
            pass
print(str(len(links)) + " links found")

latExistingShops = []
lngExistingShops = []
#add current direcorty/latlng.txt to latExistingShops and lngExistingShops
with open('/Users/ethansapp/Downloads/vapeShop/latlng.txt') as f:
    for line in f:
        latExistingShops.append(line.split(',')[0])
        lngExistingShops.append(line.split(',')[1])
print("existing shops: " + str(len(latExistingShops)))


addresses = []
rates = []
sqft = []
mile1 = []
mile3 = []
mile5 = []
mile10 = []
print("getting data")
print("links: " + str(len(links)))

def waitUntilPageLoads():
    while True:
        try:
            driver.execute_script("return document.getElementsByClassName('text')[3].innerText")
            break
        except:
            time.sleep(1)
            continue
oldVapeShops = []
with open('/Users/ethansapp/Downloads/vapeShop/oldVapeShops.txt') as f:
    for line in f:
        oldVapeShops.append(line)

for link in links:
    if link in oldVapeShops:
        continue
    else:
        with open('/Users/ethansapp/Downloads/vapeShop/oldVapeShops.txt', 'a') as f:
            f.write(str(link) + '\n')
    print(link)
    addy = link.split('/')[6]
    try:
        response = requests.get('https://maps.googleapis.com/maps/api/geocode/json?address='+addy+'&key='+googleApiKey)
        resp_json_payload = response.json()
        lat = resp_json_payload['results'][0]['geometry']['location']['lat']
        lng = resp_json_payload['results'][0]['geometry']['location']['lng']
    except:
        print("error")
        continue
    addresses.append(link.replace('\n', ''))
    count = 0
    for i in range(0, len(latExistingShops)):
        if distance.distance((lat, lng), (latExistingShops[i], lngExistingShops[i])).miles < 1:
            count += 1
    mile1.append(count)
    print("1 mile: " + str(count))
    #get number of all locations within 3 miles
    count = 0
    for i in range(0, len(latExistingShops)):
        if distance.distance((lat, lng), (latExistingShops[i], lngExistingShops[i])).miles < 3:
            count += 1
    mile3.append(count)
    print("3 miles: " + str(count))
    #get number of all locations within 5 miles
    count = 0
    for i in range(0, len(latExistingShops)):
        if distance.distance((lat, lng), (latExistingShops[i], lngExistingShops[i])).miles < 5:
            count += 1
    mile5.append(count)
    print("5 miles: " + str(count))
    #get number of all locations within 10 miles
    count = 0
    for i in range(0, len(latExistingShops)):
        if distance.distance((lat, lng), (latExistingShops[i], lngExistingShops[i])).miles < 10:
            count += 1
    mile10.append(count)
    print("10 miles: " + str(count))
    print("done")
organizedShops = []

#write to spreadsheet
print("writing to spreadsheet")
#if file doesn't exist, create it
print(len(addresses))
print(len(mile1))
print(len(mile3))
print(len(mile5))
print(len(mile10))
#check if vapeShops.xlsx exists, if not create it

if not os.path.exists("Users/ethansapp/Downloads/vapeShop/vapeShops.xlsx"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.cell(row=1, column=1).value = "Address"
    sheet.cell(row=1, column=2).value = "1 mile"
    sheet.cell(row=1, column=3).value = "3 miles"
    sheet.cell(row=1, column=4).value = "5 miles"
    sheet.cell(row=1, column=5).value = "10 miles"
    wb.save("vapeShops.xlsx")
wb = openpyxl.load_workbook("vapeShops.xlsx")
sheet = wb["Sheet1"]
#organize data
for i in range(0, len(addresses)):
    sheet.cell(row=i+2, column=1).value = addresses[i]
    sheet.cell(row=i+2, column=2).value = mile1[i]
    sheet.cell(row=i+2, column=3).value = mile3[i]
    sheet.cell(row=i+2, column=4).value = mile5[i]
    sheet.cell(row=i+2, column=5).value = mile10[i]
wb.save("vapeShops.xlsx")

#close the driver
print("closing driver")
driver.close()


